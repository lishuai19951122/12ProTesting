# 读取excel内容
from openpyxl import load_workbook

wb = load_workbook(filename='dates/empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)
for i in range(1, 31):
    print(sheet_ranges.cell(column=1, row=i).value)

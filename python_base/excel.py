# 写入excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'dates/empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

ws4 = wb.create_sheet(title="my_sheet")
for i in range(1, 41):
    ws4.cell(row=1, column=i, value='hello')

wb.save(filename=dest_filename)
